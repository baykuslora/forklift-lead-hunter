import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def export_leads_to_excel(leads: list, filename: str) -> str:
    if not leads:
        return None
    data = []
    for l in leads:
        data.append({
            "Firma Adı": l.get("company_name"),
            "Şehir / Konum": l.get("city"),
            "İlanda Çıkan Tel": l.get("direct_phone"),
            "Bulunan Santral/Merkez Tel": l.get("enriched_phone"),
            "İlan Pozisyonu": l.get("job_title"),
            "Kaynak": l.get("source"),
            "İlan Linki": l.get("job_url"),
            "Satış Arama Durumu": "Aranmadı",
            "Görüşülen Yetkili": "",
            "Satış Notları": ""
        })
    df = pd.DataFrame(data)
    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Potansiyel Leadler")
        ws = writer.sheets["Potansiyel Leadler"]
        header_fill = PatternFill(start_color="FFCC00", end_color="FFCC00", fill_type="solid")
        header_font = Font(name="Arial", size=11, bold=True, color="000000")
        thin_border = Border(
            left=Side(style='thin', color='DDDDDD'),
            right=Side(style='thin', color='DDDDDD'),
            top=Side(style='thin', color='DDDDDD'),
            bottom=Side(style='thin', color='DDDDDD')
        )
        for col_idx, col in enumerate(ws.iter_cols(min_row=1, max_row=len(df)+1), 1):
            col[0].fill = header_fill
            col[0].font = header_font
            col[0].alignment = Alignment(horizontal="center", vertical="center")
            for cell in col[1:]:
                cell.border = thin_border
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[get_column_letter(col_idx)].width = max(max_len + 4, 15)
        ws.row_dimensions[1].height = 26
        ws.freeze_panes = "A2"
    return filename